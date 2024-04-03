import numpy as np
import matplotlib.pyplot as plt
from tqdm import tqdm
import openpyxl
from time import time
from iva_g import iva_g
from helpers_iva import whiten_data
import cProfile
from titan_iva_g_problem_simulation import *
from titan_iva_g_reg import *
from titan_iva_g_lab import *



def search_criteria(algo='palm_iva_g_reg',K=5,N=3,T=10000,max_iter=5000,criteria_max=10**(-10),rhos=[0.4,0.6],lambda_=0.1,gamma_c=1.9,gamma_w=0.9,alpha=10,seed=1):
    Sigma = make_Sigma(K,N,rank=K+10,mu=rhos,lambda_=lambda_,seed=999,normalize=False)
    S = make_S(Sigma,T)
    A = make_A(K,N)
    X = make_X(S,A)
    X_,U = whiten_data(X)
    A_ = np.einsum('nNk, Nvk-> nvk', U, A)
    if algo == 'palm_iva_g_reg':
        _,_,_,isi,diffs = palm_iva_g_reg(X_,alpha=alpha,max_iter=max_iter,criteria=criteria_max,track_isi=True,track_diff=True,B=A_,seed=seed,update_scheme=(1,5),gamma_w=gamma_w,gamma_c=gamma_c,inertial=False)
    elif algo == 'newton':
        _,_,_,isi,diffs = iva_g(X_,opt_approach='newton',A=A_,max_iter=max_iter,W_diff_stop=criteria_max,return_W_change=True)
    elif algo == 'gradient':
        _,_,_,isi,diffs = iva_g(X_,opt_approach='gradient',A=A_,max_iter=max_iter,W_diff_stop=criteria_max,return_W_change=True)
    else:
        raise('You must provide one of the following algorithms : palm_iva_g_reg, newton or gradient')
    # diffs_max = []
    # diffs_W,diffs_C = diffs
    # for i in range(len(diffs_W)):
    #     diffs_max.append(max(diffs_W[i],diffs_C[i]))
    diffs_max = diffs
    thresholds = []
    index = 0
    power = 0
    while index < len(diffs_max):
        if diffs_max[index] < 10**(-power):
            thresholds.append(index)
            power += 1
        else:
            index += 1
    fig,ax1 = plt.subplots()
    ax1.set_title('ISI score for various stopping criteria. K = {}, N = {}'.format(K,N), fontsize=16,y=1.05,x=0.5)
    ax1.plot(isi,color='y',label='ISI(i))')   
    for i,threshold in enumerate(thresholds):
        if i > 6:
            ax1.axvline(x=threshold,color='r')
            ax1.text(threshold,ax1.get_ylim()[1]+0.6,'e-{}'.format(i),color='k',ha='center',fontsize=14) #va='bottom'
    ax1.set_xlabel('iteration $i$',fontsize=16)
    ax1.set_ylabel('ISI score',fontsize=16)
    ax1.set_yscale('log')
    ax1.set_ylim(1e-4,1e0)
    ax1.legend(loc=1)
    plt.show()
            

# compare_alphas(K=5,N=5,seed=2)
search_criteria(max_iter=20000,rhos=[0.2,0.3],lambda_=0.04,K=5,N=10,alpha=1)

def compare_diffs(N_exp=1,K=5,N=3,T=10000,rhos=[0.6,0.7],lambda_=0.04,max_iter=20000,criteria=10**(-8),alpha=10,gamma_w=0.9,gamma_c=1.9,seed=None):
    for exp in tqdm(range(N_exp)):
        Sigma = make_Sigma(K,N,rank=K+10,mu=rhos,lambda_=lambda_,seed=None,normalize=False)
        S = make_S(Sigma,T)
        A = make_A(K,N)
        X = make_X(S,A)
        X_,U = whiten_data(X)
        A_ = np.einsum('nNk, Nvk-> nvk',U,A)
        _,_,_,_,diffs = palm_iva_g_reg(X_,gamma_c=gamma_c,gamma_w=gamma_w,alpha=alpha,max_iter=max_iter,criteria=criteria,B=A_,track_diff=True,seed=seed)
        # _,_,_,_,diffs = titan_palm_iva_g_reg(X_,gamma_c=1,gamma_w=gamma_w,alpha=alpha,max_iter=max_iter,criteria=criteria,track_diff=True,seed=seed)
        diffs_W,diffs_C = diffs
        fig,ax1 = plt.subplots()
        ax1.set_title('evolution of the criterias with repect to W (blue) and C (red). K = {}, N = {}'.format(K,N),fontsize=16,y=1.05,x=0.5)
        ax1.set_ylabel('criteria',fontsize=16)
        ax1.set_xlabel('iteration $i$',fontsize=16)
        ax1.set_yscale('log')
        ax1.plot(diffs_W,color ='b',label='$theta_W^{(i)}$')
        ax1.plot(diffs_C,color ='r',label='$theta_C^{(i)}$')
        ax1.legend(loc=1)
        plt.show()
        

# compare_diffs(N_exp=1,K=5,N=3,criteria=10**(-8),max_iter=20000,lambda_=0.1,seed=1,alpha=1)

def cost_evolution(N_exp=1,K=5,N=3,T=10000,rhos=[0.6,0.7],lambda_=0.04,max_iter=10000,criteria=10**(-12),alpha=10,gamma_w=0.9,gamma_c=1.9,seed=None):
    wb = openpyxl.Workbook()
    date = time()
    file_name = "evolution of cost for updates of W and C for K = {} and N = {} and lambda = {} and rho = {} at date {}.xlsx".format(K,N,lambda_,(rhos[0]+rhos[1])/2,date)
    sheet = wb.active
    for exp in tqdm(range(N_exp)):
        Sigma = make_Sigma(K,N,rank=K+10,mu=rhos,lambda_=lambda_,seed=None,normalize=False)
        S = make_S(Sigma,T)
        A = make_A(K,N)
        X = make_X(S,A)
        X_,U = whiten_data(X)
        A_ = np.einsum('nNk, Nvk-> nvk',U,A)
        # _,_,_,cost,_,_ = palm_iva_g_reg(X_,gamma_c=gamma_c,gamma_w=gamma_w,alpha=alpha,max_iter=max_iter,criteria=criteria,B=A_,track_diff=True,seed=seed)
        _,_,_,cost = titan_iva_g_reg(X_,gamma_c=1,gamma_w=gamma_w,alpha=alpha,max_iter=max_iter,update_scheme=(20,1),criteria=criteria,B=A_,track_cost=True,seed=seed)
        cost_W = []
        cost_C = []
        for i in range(len(cost)//2):
            cost_W.append(cost[2*i+1]-cost[2*i])
            cost_C.append(cost[2*i+2]-cost[2*i+1])
        if exp > 0:
            sheet = wb.create_sheet("Feuille {}".format(exp+1))
            for index, value in enumerate(cost_W,1):
                sheet.cell(row=index, column=3, value=value)
            for index, value in enumerate(cost_C,1):
                sheet.cell(row=index, column=4, value=value)
        wb.save(file_name)
        # print(f'Le fichier Excel "{file_name}" a été créé avec succès.')
        fig,ax1 = plt.subplots()
        ax1.set_title('evolution of the cost when we update W (yellow) and C(green). K = {}, N = {}'.format(K,N), fontsize=16,y=1.05,x=0.5)
        ax1.plot(cost_W,color='y',label='block W')
        ax1.plot(cost_C,color='g',label='block C')
        ax1.axhline(y=0,color='k')
        l = 10*max(np.abs(cost_C[-1]),np.abs(cost_W[-1]))
        ax1.set_ylim(-l,l)
        ax1.set_xlabel('iteration $i$',fontsize=16)
        ax1.set_ylabel('cost evolution',fontsize=16)
        ax1.legend(loc=1)
        print(decrease(cost))
        plt.show()

# cost_evolution(N_exp=1,seed=5)

# [0.1,0.5,1,5,10,20,50,100]
def search_alphas_and_criterias(K=5,N=5,T=10000,max_iter=20000,criteria_max=10**(-15),alphas=[0.1,0.5,1,5,10,20,50,100],seed=None,rhos=[0.6,0.7],lambda_=0.04,gamma_c=0.6,gamma_w=1.1):
    Sigma = make_Sigma(K,N,rank=K+10,mu=rhos,lambda_=lambda_,seed=999,normalize=False)
    S = make_S(Sigma,T)
    A = make_A(K,N)
    X = make_X(S,A)
    X_,U = whiten_data(X)
    A_ = np.einsum('nNk, Nvk-> nvk', U, A)
    fig,ax1 = plt.subplots()
    for alpha in alphas:
        print('coucou')
        _,_,_,cost_WC,isi,diffs_WC = palm_iva_g_reg(X_,alpha=alpha,max_iter=max_iter,criteria=criteria_max,track_diff=True,B=A_,seed=seed,gamma_w=gamma_w,gamma_c=gamma_c)
        c = np.random.uniform(0.5,1,3)
        ax1.plot(isi,label='alpha = {}'.format(alpha),color=c)
        ax1.legend(loc=1)
        diffs_W,diffs_C = diffs_WC
        diffs = []
        cost = []
        for i in range(len(diffs_W)):
            diffs.append(max(diffs_W[i],diffs_C[i]))
            cost.append(cost_WC[2*i])
        thresholds = []
        index = 0
        power = 0
        while index < len(diffs):
            if diffs[index] < 10**(-power):
                thresholds.append(index)
                power += 1
            else:
                index += 1
        for i,threshold in enumerate(thresholds):
            if i > 6:
                ax1.plot(threshold,isi[threshold], marker="o",color=c)
                ax1.text(threshold,isi[threshold]+0.01,'e-{}'.format(i),color='k',ha='center',fontsize=14) #va='bottom'
    ax1.set_title('ISI curve for several values of alpha, K = {}, N = {}'.format(K,N), fontsize=16, loc = 'right')
    ax1.set_yscale('log')
    ax1.set_xlabel('iteration $i$',fontsize=16)
    ax1.set_ylabel('ISI score',fontsize=16)
    plt.show()

# search_alphas_and_criterias()

# K = 5
# N = 5
# mu = [0.6,0.7]
# lambdas = [0.01,0.04,0.1,0.16,0.25]

# # def identifiability_law(N=3,K=2,N_exp=100,mu=mu):
# #     fig,ax = plt.subplots()


# ident = np.zeros(100)
# for seed in tqdm(range(100)):
#     Sigma = make_Sigma_2(K,N,K+10,mu,lambdas[4],seed=seed)
#     ident[seed] = identifiability_level(Sigma)
# print(np.mean(ident))
# print(np.std(ident))
# print(np.min(ident))